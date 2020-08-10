#!/usr/bin/env python3

"""Report on related articles.

From the ticket:

  In order to get a better understanding of the breadth and volume of
  related citations to inform requirements for OCEEBMS-568, would it
  be possible to generate an ad-hoc spreadsheet showing related
  citations for all those articles imported into the EBMS since Jan 1,
  2020? It would be helpful if it's possible for the spreadsheet to
  include the following information:

    * PMID of the citation in the EBMS
    * the journal abbreviation for the citation in the EBMS
    * the date of publication of the citation
    * the date the citation was imported into the EBMS
    * the PMID of the related citation
    * the type of related citation (if this is something PubMed provides?
      It would be something like editorial/comment, supplement, errata)
    * the journal abbreviation for the related citation
    * the date of publication of the related citation

Example block from PubMed XML:

        <CommentsCorrectionsList>
            <CommentsCorrections RefType="CommentOn">
                <RefSource>Lancet Oncol. 2020 Jan;21(1):105-120</RefSource>
                <PMID Version="1">31753727</PMID>
            </CommentsCorrections>
        </CommentsCorrectionsList>
"""

import argparse
import datetime
import getpass
import os
import sys
import pymysql
import openpyxl
from lxml import etree

SQL = "SELECT article_id FROM ebms_article WHERE import_date >= '2020-01-01'"
PATH = "MedlineCitation/Article/PublicationTypeList/PublicationType"
PATH = "MedlineCitation/CommentsCorrectionsList/CommentsCorrections"
COLS = (
    "PMID",
    "Journal",
    "Published",
    "Imported",
    "Related",
    "Ref Type",
    "Ref Source",
)

class CommentsCorrections:
    def __init__(self, node):
        self.node = node
    @property
    def ref_source(self):
        node = self.node.find("RefSource")
        return node.text if node is not None else None
    @property
    def pmid(self):
        node = self.node.find("PMID")
        return node.text if node is not None else None
    @property
    def ref_type(self):
        return self.node.get("RefType")

parser = argparse.ArgumentParser()
parser.add_argument("--host", required=True)
parser.add_argument("--port", type=int, default=3661)
parser.add_argument("--db", default="oce_ebms")
parser.add_argument("--user", default="oce_ebms")
opts = vars(parser.parse_args())
opts["passwd"] = getpass.getpass("password for %s: " % opts["user"])
cursor = pymysql.connect(**opts).cursor()
cursor.execute("SET NAMES utf8")
cursor.execute(SQL)
ids = [row[0] for row in cursor.fetchall()]
rows = []
done = 0
for id in ids:
    cursor.execute("""\
SELECT source_id,
       brf_jrnl_title,
       published_date,
       import_date,
       source_data
  FROM ebms_article
 WHERE article_id = %s""", (id,))
    pmid, journal, published, imported, xml = cursor.fetchone()
    root = etree.fromstring(xml.encode("utf-8"))
    for node in root.findall(PATH):
        cc = CommentsCorrections(node)
        imported = str(imported)[:10]
        rows.append((pmid, journal, published, imported,
                     cc.pmid, cc.ref_type, cc.ref_source))
    row = cursor.fetchone()
    done += 1
    sys.stderr.write(f"\rparsed {done} of {len(ids)} articles")
book = openpyxl.Workbook()
sheet = book.active
sheet.title = "Related"
for c, name in enumerate(COLS):
    sheet.cell(row=1, column=c+1, value=name)
r = 2
for row in rows:
    for c, value in enumerate(row):
        sheet.cell(row=r, column=c+1, value=value)
    r += 1
book.save("oceebms-570.xlsx")
sys.stderr.write("\nwrote oceebms-570.xlsx\n")
